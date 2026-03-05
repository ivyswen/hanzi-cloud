import os
import re
import collections
from typing import List, Dict

import jieba
import docx
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QFileDialog, QTableWidget, QTableWidgetItem,
    QLabel, QProgressBar, QMessageBox, QHeaderView,
    QDialog, QDialogButtonBox, QPlainTextEdit
)
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QPixmap, QImage
from wordcloud import WordCloud
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from stopwords import load_stopwords, save_stopwords

class AnalysisWorker(QThread):
    """后台处理线程，防止GUI卡死"""
    progress = Signal(int)
    finished = Signal(dict, object)  # 返回统计结果和词云字节流
    error = Signal(str)

    def __init__(self, file_paths: List[str], stopwords: set):
        super().__init__()
        self.file_paths = file_paths
        self.stopwords = stopwords   # 从主窗口传入，支持运行时更新

    def run(self):
        all_text = ""
        total_files = len(self.file_paths)
        
        try:
            for i, path in enumerate(self.file_paths):
                # 1. 读取 Word 文档
                doc = docx.Document(path)
                text = "\n".join([para.text for para in doc.paragraphs])
                all_text += text
                self.progress.emit(int(((i + 1) / total_files) * 50))

            # 2. 清洗数据（去除特殊符号）
            all_text = re.sub(r'[^\u4e00-\u9fa5]+', ' ', all_text)

            # 3. 分词
            words = jieba.lcut(all_text)
            filtered_words = [w for w in words if len(w) > 1 and w not in self.stopwords]
            
            self.progress.emit(70)

            # 4. 统计词频
            word_counts = collections.Counter(filtered_words)
            top_words = dict(word_counts.most_common(100))

            # 5. 生成词云图
            # 注意：这里需要系统内有中文字体，否则会乱码。
            # Windows通常路径: C:\Windows\Fonts\simhei.ttf
            font_path = "simhei.ttf" if os.name == 'nt' else "/System/Library/Fonts/STHeiti Light.ttc"
            
            wc = WordCloud(
                font_path=font_path,
                background_color='white',
                width=800,
                height=600,
                max_words=100
            ).generate_from_frequencies(top_words)

            # 使用非交互式后端将词云渲染到内存，避免在子线程中触发 GUI
            img_buffer = BytesIO()
            fig = Figure(figsize=(8, 6))
            FigureCanvasAgg(fig)          # 绑定 Agg 后端（无 GUI）
            ax = fig.add_subplot(111)
            ax.imshow(wc, interpolation='bilinear')
            ax.axis('off')
            fig.savefig(img_buffer, format='png', bbox_inches='tight', pad_inches=0)
            img_buffer.seek(0)

            self.progress.emit(100)
            self.finished.emit(top_words, img_buffer.read())

        except Exception as e:
            self.error.emit(str(e))

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("中文文档批量词频分析工具")
        self.resize(1100, 750)
        self.selected_files = []
        self.word_data: dict = {}       # 缓存最新一次分析结果，供导出使用
        self.stopwords: set = load_stopwords()  # 启动时从文件加载停用词
        self.init_ui()

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)

        # --- 左侧控制与列表面板 ---
        left_panel = QVBoxLayout()
        
        self.btn_select = QPushButton("选择 Word 文档 (.docx)")
        self.btn_select.setFixedHeight(40)
        self.btn_select.clicked.connect(self.select_files)
        left_panel.addWidget(self.btn_select)

        self.lbl_info = QLabel("未选择文件")
        left_panel.addWidget(self.lbl_info)

        self.btn_run = QPushButton("开始分析")
        self.btn_run.setFixedHeight(40)
        self.btn_run.setEnabled(False)
        self.btn_run.clicked.connect(self.start_analysis)
        left_panel.addWidget(self.btn_run)

        self.btn_stopwords = QPushButton("管理停用词")
        self.btn_stopwords.setFixedHeight(40)
        self.btn_stopwords.clicked.connect(self.open_stopwords_dialog)
        left_panel.addWidget(self.btn_stopwords)

        self.progress_bar = QProgressBar()
        left_panel.addWidget(self.progress_bar)

        self.btn_export = QPushButton("导出 Excel")
        self.btn_export.setFixedHeight(40)
        self.btn_export.setEnabled(False)   # 分析完成前保持禁用
        self.btn_export.clicked.connect(self.export_excel)
        left_panel.addWidget(self.btn_export)

        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["词语", "频次"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        left_panel.addWidget(self.table)

        main_layout.addLayout(left_panel, 1)

        # --- 右侧可视化面板 ---
        self.right_panel = QVBoxLayout()
        self.lbl_cloud = QLabel("词云图展示区域")
        self.lbl_cloud.setAlignment(Qt.AlignCenter)
        self.lbl_cloud.setStyleSheet("border: 1px dashed #ccc; background: #f9f9f9;")
        self.right_panel.addWidget(self.lbl_cloud)
        
        main_layout.addLayout(self.right_panel, 2)

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "选择 Word 文档", "", "Word Documents (*.docx)"
        )
        if files:
            self.selected_files = files
            self.lbl_info.setText(f"已选择 {len(files)} 个文件")
            self.btn_run.setEnabled(True)

    def start_analysis(self):
        self.btn_run.setEnabled(False)
        self.table.setRowCount(0)
        self.progress_bar.setValue(0)
        
        self.worker = AnalysisWorker(self.selected_files, self.stopwords)
        self.worker.progress.connect(self.progress_bar.setValue)
        self.worker.error.connect(self.handle_error)
        self.worker.finished.connect(self.display_results)
        self.worker.start()

    def handle_error(self, err_msg):
        QMessageBox.critical(self, "错误", f"分析过程中出现问题：\n{err_msg}")
        self.btn_run.setEnabled(True)

    def display_results(self, word_dict, img_data):
        # 1. 更新表格
        self.table.setRowCount(len(word_dict))
        for i, (word, count) in enumerate(word_dict.items()):
            self.table.setItem(i, 0, QTableWidgetItem(word))
            self.table.setItem(i, 1, QTableWidgetItem(str(count)))

        # 2. 更新词云图
        pixmap = QPixmap()
        pixmap.loadFromData(img_data)
        
        # 适应窗口大小缩放
        scaled_pixmap = pixmap.scaled(
            self.lbl_cloud.width() - 20, 
            self.lbl_cloud.height() - 20, 
            Qt.KeepAspectRatio, 
            Qt.SmoothTransformation
        )
        self.lbl_cloud.setPixmap(scaled_pixmap)
        self.btn_run.setEnabled(True)

        # 缓存结果并解锁导出按钮
        self.word_data = word_dict
        self.btn_export.setEnabled(True)

    def export_excel(self):
        """将词频统计结果导出为带样式的 Excel 文件"""
        path, _ = QFileDialog.getSaveFileName(
            self, "保存 Excel 文件", "词频统计.xlsx", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "词频统计"

            # 写入表头（白字蓝底加粗居中）
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4472C4")
            for col, text in enumerate(["词语", "频次"], start=1):
                cell = ws.cell(row=1, column=col, value=text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # 写入数据行
            for row, (word, count) in enumerate(self.word_data.items(), start=2):
                ws.cell(row=row, column=1, value=word)
                ws.cell(row=row, column=2, value=count)

            # 设置列宽
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 10

            wb.save(path)
            QMessageBox.information(self, "导出成功", f"文件已保存至：\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

    def open_stopwords_dialog(self):
        """打开停用词管理对话框"""
        dlg = StopwordsDialog(self.stopwords, self)
        if dlg.exec() == QDialog.Accepted:
            # 用户点击保存：更新内存中的停用词集合
            self.stopwords = dlg.get_stopwords()
            save_stopwords(self.stopwords)
            QMessageBox.information(
                self, "已保存",
                f"停用词已更新（共 {len(self.stopwords)} 个），下次分析生效。"
            )


class StopwordsDialog(QDialog):
    """停用词管理对话框：支持查看、增删词并持久化"""

    def __init__(self, stopwords: set, parent=None):
        super().__init__(parent)
        self.setWindowTitle("管理停用词")
        self.resize(400, 500)
        self._default_stopwords = set(stopwords)   # 保留原始副本供重置用

        layout = QVBoxLayout(self)

        # 说明标签
        hint = QLabel("每行一个停用词，保存后下次分析生效：")
        layout.addWidget(hint)

        # 文本编辑区：展示当前停用词（按字典序）
        self.editor = QPlainTextEdit()
        self.editor.setPlaceholderText("在此输入停用词，每行一个…")
        self.editor.setPlainText("\n".join(sorted(stopwords)))
        layout.addWidget(self.editor)

        # 底部按钮区：重置默认 / Cancel / OK
        btn_box = QDialogButtonBox()
        btn_reset = QPushButton("重置为默认")
        btn_reset.clicked.connect(self._reset_to_default)
        btn_box.addButton(btn_reset, QDialogButtonBox.ResetRole)
        btn_box.addButton(QDialogButtonBox.Cancel)
        btn_box.addButton(QDialogButtonBox.Ok)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def _reset_to_default(self):
        """将编辑器内容重置为打开时的原始词集"""
        self.editor.setPlainText("\n".join(sorted(self._default_stopwords)))

    def get_stopwords(self) -> set:
        """从编辑器解析停用词集合（自动过滤空行）"""
        lines = self.editor.toPlainText().splitlines()
        return {line.strip() for line in lines if line.strip()}


if __name__ == "__main__":
    app = QApplication([])
    window = MainWindow()
    window.show()
    app.exec()